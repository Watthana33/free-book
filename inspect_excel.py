import openpyxl
import json
import os

excel_path = r"C:\Users\Acer\Desktop\free_book\รายการสั่งหนังสือ วิทยาลัยเทคนิคอุดรธานี2569.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

print("Sheet names in Workbook:", wb.sheetnames)

summary_data = {}

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    max_r = sheet.max_row
    max_c = sheet.max_column
    
    rows_sample = []
    non_empty_count = 0
    for r in range(1, max_r + 1):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, max_c + 1)]
        # clean row_vals
        row_str_vals = [str(v) if v is not None else "" for v in row_vals]
        if any(v != "" for v in row_str_vals):
            non_empty_count += 1
            if len(rows_sample) < 25:
                rows_sample.append({f"R{r}": row_str_vals[:15]})
                
    summary_data[sheet_name] = {
        "max_row": max_r,
        "max_col": max_c,
        "non_empty_rows": non_empty_count,
        "sample_rows": rows_sample
    }

with open("excel_structure.json", "w", encoding="utf-8") as f:
    json.dump(summary_data, f, ensure_ascii=False, indent=2)

print("Excel structure exported to excel_structure.json")
