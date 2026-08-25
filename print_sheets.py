import openpyxl
import io
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\Acer\Desktop\free_book\รายการสั่งหนังสือ วิทยาลัยเทคนิคอุดรธานี2569.xlsx', data_only=True)
print(f"Total Sheets: {len(wb.sheetnames)}")

for i, name in enumerate(wb.sheetnames):
    s = wb[name]
    print(f"\n=========================================")
    print(f"Sheet {i}: {name} (max_row={s.max_row}, max_col={s.max_column})")
    print(f"=========================================")
    row_count = 0
    for r in range(1, s.max_row + 1):
        vals = [s.cell(row=r, column=c).value for c in range(1, s.max_column + 1)]
        if any(v is not None and str(v).strip() != '' for v in vals):
            row_count += 1
            if row_count <= 30:
                non_empty = [f"C{c}:{v}" for c, v in enumerate(vals, 1) if v is not None and str(v).strip() != '']
                print(f"Row {r:3d}: " + " | ".join(non_empty[:10]))
    print(f"Total non-empty rows: {row_count}")
