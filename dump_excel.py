import openpyxl

excel_path = r"C:\Users\Acer\Desktop\free_book\รายการสั่งหนังสือ วิทยาลัยเทคนิคอุดรธานี2569.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

with open("full_dump.txt", "w", encoding="utf-8") as out:
    for idx, sheet_name in enumerate(wb.sheetnames):
        sheet = wb[sheet_name]
        out.write(f"\n=========================================\n")
        out.write(f"SHEET Index {idx}: raw_name = {repr(sheet_name)}, max_row={sheet.max_row}, max_col={sheet.max_column}\n")
        out.write(f"=========================================\n")
        
        for r in range(1, sheet.max_row + 1):
            row_vals = []
            for c in range(1, sheet.max_column + 1):
                val = sheet.cell(row=r, column=c).value
                if val is not None:
                    row_vals.append(f"C{c}: {repr(val)}")
            if row_vals:
                out.write(f"Row {r:3d}: " + " | ".join(row_vals) + "\n")

print("Full dump written to full_dump.txt")
